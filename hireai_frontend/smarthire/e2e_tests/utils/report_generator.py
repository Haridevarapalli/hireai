import os
import json
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def generate_reports():
    # Paths
    results_dir = os.path.join(os.getcwd(), "Test Results")
    os.makedirs(os.path.join(results_dir, "Excel"), exist_ok=True)
    os.makedirs(os.path.join(results_dir, "Summary"), exist_ok=True)
    
    excel_path = os.path.join(results_dir, "Excel", "Automation_Test_Report.xlsx")
    summary_path = os.path.join(results_dir, "Summary", "summary.md")
    xml_path = os.path.join(results_dir, "report.xml") # Assuming pytest --junitxml was used
    
    # Check if XML exists
    if not os.path.exists(xml_path):
        print("JUnit XML report not found. Cannot generate Excel/Markdown summaries.")
        return
        
    tree = ET.parse(xml_path)
    root = tree.getroot()
    
    total = int(root.attrib.get('tests', 0))
    errors = int(root.attrib.get('errors', 0))
    failures = int(root.attrib.get('failures', 0))
    skipped = int(root.attrib.get('skipped', 0))
    passed = total - errors - failures - skipped
    
    pass_percentage = (passed / total * 100) if total > 0 else 0
    
    # 1. Generate Excel Report
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"
    
    # Headers
    headers = ["Test Case", "Class Name", "Status", "Time (s)", "Message"]
    ws.append(headers)
    
    # Style Headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
    failed_tests = []
    
    for testcase in root.findall('.//testcase'):
        name = testcase.attrib.get('name', '')
        classname = testcase.attrib.get('classname', '')
        time = testcase.attrib.get('time', '0')
        
        status = "Passed"
        message = ""
        
        if testcase.find('failure') is not None:
            status = "Failed"
            message = testcase.find('failure').get('message', '')
            failed_tests.append({"name": name, "reason": message})
        elif testcase.find('error') is not None:
            status = "Error"
            message = testcase.find('error').get('message', '')
            failed_tests.append({"name": name, "reason": message})
        elif testcase.find('skipped') is not None:
            status = "Skipped"
            message = testcase.find('skipped').get('message', '')
            
        ws.append([name, classname, status, time, message])
        
    wb.save(excel_path)
    
    # 2. Generate Markdown Summary
    base_url = os.environ.get("BASE_URL", "https://<github-username>.github.io/<repository-name>/")
    
    md_content = f"""# Live GitHub Pages E2E Test Summary

**Deployment URL:**
{base_url}

**Total Tests:** {total}
**Passed:** {passed}
**Failed:** {failures + errors}
**Skipped:** {skipped}
**Pass Percentage:** {pass_percentage:.2f}%

"""
    if failed_tests:
        md_content += "### Failed Tests:\n"
        for ft in failed_tests:
            reason = ft['reason'].split('\\n')[0] if ft['reason'] else "Unknown"
            md_content += f"- **{ft['name']}**: {reason}\n"
    else:
        md_content += "### Failed Tests:\n- None! All tests passed.\n"
        
    with open(summary_path, "w") as f:
        f.write(md_content)
        
    print(f"Reports generated successfully at {results_dir}")

if __name__ == "__main__":
    generate_reports()
